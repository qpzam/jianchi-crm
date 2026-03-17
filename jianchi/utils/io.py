"""
文件读写工具集
整合自: match_contacts, generate_worklist, auto_fetch_and_match
"""
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


def load_dataframe(filepath: str, **kwargs) -> pd.DataFrame:
    """
    智能加载 Excel / CSV / JSON / TXT，自动处理编码
    """
    path = Path(filepath)
    ext = path.suffix.lower()

    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(filepath, dtype=str, **kwargs)
    elif ext == ".csv":
        for enc in ("utf-8", "gbk", "gb2312", "gb18030"):
            try:
                df = pd.read_csv(filepath, dtype=str, encoding=enc, **kwargs)
                break
            except (UnicodeDecodeError, Exception):
                continue
        else:
            raise ValueError(f"无法识别文件编码: {filepath}")
    elif ext == ".json":
        with open(filepath, "r", encoding="utf-8") as f:
            df = pd.DataFrame(json.load(f))
    elif ext == ".txt":
        df = _load_contact_txt(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")

    df.columns = df.columns.str.strip()
    return df


def _load_contact_txt(filepath: str) -> pd.DataFrame:
    """解析 txt 格式联系方式库（每行: 公司名 联系人 [职务] 电话：xxx）"""
    records = []
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            phone_match = re.search(r'电话[:：]\s*(\d{11})', line)
            if not phone_match:
                continue
            parts = line.split()
            if len(parts) < 2:
                continue
            records.append({
                "公司": parts[0],
                "联系人": parts[1] if len(parts) > 1 else "",
                "职务": parts[2] if len(parts) > 2 and "电话" not in parts[2] else "",
                "手机": phone_match.group(1),
            })
    return pd.DataFrame(records)


def auto_map_columns(df: pd.DataFrame, candidates: dict) -> dict:
    """
    自动识别并映射列名
    candidates = {"standard_name": ["候选1", "候选2", ...]}
    返回 {"standard_name": "实际列名"}
    """
    result = {}
    df_cols_lower = {c.strip().lower(): c for c in df.columns}

    for standard, names in candidates.items():
        for name in names:
            if name.strip().lower() in df_cols_lower:
                result[standard] = df_cols_lower[name.strip().lower()]
                break

    return result


def save_excel(records: list[dict], filepath: str, sheet_name: str = "Sheet1",
               color_rules: dict | None = None):
    """
    保存记录列表到 Excel，带基本格式

    color_rules 示例:
    {
        "column_index": 2,
        "rules": [
            {"contains": ["有限公司", "基金"], "color": "FF9900"},
            {"default": True, "color": "CC0000"},
        ]
    }
    """
    from openpyxl import Workbook

    if not records:
        print("⚠️ 无数据可保存")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头
    columns = list(records[0].keys())
    ws.append(columns)

    # 表头样式
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for rec in records:
        ws.append([rec.get(col, "") for col in columns])

    # 颜色规则
    if color_rules:
        col_idx = color_rules["column_index"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[col_idx]
            value = str(cell.value or "")
            for rule in color_rules["rules"]:
                if "contains" in rule:
                    if any(kw in value for kw in rule["contains"]):
                        cell.font = Font(color=rule["color"])
                        break
                elif rule.get("default"):
                    cell.font = Font(color=rule["color"])

    # 自动列宽
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(filepath)
    print(f"✓ 已保存: {filepath} ({len(records)} 条)")
